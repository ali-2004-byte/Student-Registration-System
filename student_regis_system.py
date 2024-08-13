from tkinter import * 
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl.workbook
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib


root=Tk()
root.title("Cityzen Registration App")
root.geometry("1250x700+300+200")
root.configure(bg="#202329")
root.resizable(True,True)

#logo = PhotoImage(file=file_path+"mcfc.png")
#root.iconphoto(True,logo)

file=pathlib.Path('People_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet = file.active
    sheet['A1']="Registration No"
    sheet['B1']="Date of Registration"
    sheet['C1']="Name"
    sheet['D1']="DOB"
    sheet['E1']="Gender"
    sheet['F1']="Class"
    sheet['G1']="Religion"
    sheet['H1']="Skills"
    sheet['I1']="Father"
    sheet['J1']="Mother"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"
    file.save('People_data.xlsx')

def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file",filetype=(("JPG File","*.jpg"),("PNG File","*.png"),("All Files","*.txt")))
    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

def registration_no():
    file=openpyxl.load_workbook("People_data.xlsx")
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value
    try:
        Registration.set(int(max_row_value)+1)
    except:
        Registration.set("1")

def clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skills.set('')
    Father.set('')
    Mother.set('')
    fo.set('')
    mo.set('')
    Class.set("Select Class")

    registration_no()
    SaveButton.config(state='normal')
    img1=PhotoImage(file='upload photo.png')
    lbl.config(image=img1)
    lbl.image=img1
    img=""

def save():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")
    D1=Date.get()
    D2=DOB.get()
    R2=Religion.get()
    S1=Skills.get()
    F1=Father.get()
    M1=Mother.get()
    F2=fo.get()
    M2=mo.get()
    if N1=="" or C1=="Select Class" or D2=="" or R2=="" or S1=="" or F1=="" or M1=="" or F2=="" or M2=="":
        messagebox.showerror("error","Few data is missing!")
    else:
        file=openpyxl.load_workbook('People_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=D1)
        sheet.cell(column=3,row=sheet.max_row,value=N1)
        sheet.cell(column=4,row=sheet.max_row,value=D2)
        sheet.cell(column=5,row=sheet.max_row,value=G1)
        sheet.cell(column=6,row=sheet.max_row,value=C1)
        sheet.cell(column=7,row=sheet.max_row,value=R2)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=F1)
        sheet.cell(column=10,row=sheet.max_row,value=M1)
        sheet.cell(column=11,row=sheet.max_row,value=F2)
        sheet.cell(column=12,row=sheet.max_row,value=M2)
        file.save(r'People_data.xlsx')
        try:
            img.save("Students/"+str(R1)+".jpg")
        except:
            messagebox.showerror("info","Profile picture is not available")

        messagebox.showinfo("info","Successfully data entered")
        clear()
        registration_no()




def exit():
    root.destroy()

def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
        print(gender)
    elif value==2:
        gender="Female"
        print(gender)


def search():
    text=Search.get()
    clear()
    SaveButton.config(state='disabled')
    file=openpyxl.load_workbook('People_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]


    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number")

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    Registration.set(x1)
    Date.set(x2)
    Name.set(x3)
    DOB.set(x4)
    radio.set(x5)
    Class.set(x6)
    Religion.set(x7)
    Skills.set(x8)
    Father.set(x9)
    Mother.set(x10)
    fo.set(x11)
    mo.set(x12)

    Img=(Image.open("Students/"+str(x1)+".jpg"))
    resized_image=Img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

def update():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")
    D1=Date.get()
    D2=DOB.get()
    R2=Religion.get()
    S1=Skills.get()
    F1=Father.get()
    M1=Mother.get()
    F2=fo.get()
    M2=mo.get()

    file=openpyxl.load_workbook("People_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    sheet.cell(column=2,row=int(reg_number),value=D1)
    sheet.cell(column=3,row=int(reg_number),value=N1)
    sheet.cell(column=4,row=int(reg_number),value=D2)
    sheet.cell(column=5,row=int(reg_number),value=G1)
    sheet.cell(column=6,row=int(reg_number),value=C1)
    sheet.cell(column=7,row=int(reg_number),value=R2)
    sheet.cell(column=8,row=int(reg_number),value=S1)
    sheet.cell(column=9,row=int(reg_number),value=F1)
    sheet.cell(column=10,row=int(reg_number),value=M1)
    sheet.cell(column=11,row=int(reg_number),value=F2)
    sheet.cell(column=12,row=int(reg_number),value=M2)

    file.save(r'People_data.xlsx')
    try:
        img.save("Students/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("Update","Updated Successfully")
    clear()



#top layer
frame1=Frame(root, width=1250,height=50,bg="#42d7fe")
frame1.place(x=0,y=0)
label1=Label(frame1,text="Email: ali_hassan2004@outlook.com",font=("baskerville",10,"bold"),bg="#42d7fe",fg="#000000")
label1.place(x=950,y=20)
#Title
frame2=Frame(root,width=1250,height=75,bg="#38b6ff")
frame2.place(x=0,y=50)
label2=Label(frame2,text="Student Registration",font=("Baskerville", 30, "bold"), bg="#38b6ff",fg="#ffffff")
label2.place(x=400,y=15)
Search=StringVar()
entry1=Entry(frame2,width=10,bd=3,textvariable=Search,font=("baskerville",20,"bold"))
entry1.place(x=830,y=20)
search_image=PhotoImage(file="search.png")
button1=Button(frame2,text="Search",compound=LEFT,image=search_image,bg="#17a9d2",width=123,font=("Baskerville",16,"bold"),fg=("#ffffff"),command=search)
button1.place(x=1000,y=20)
update_image=PhotoImage(file="Layer 4.png")
button2=Button(frame2,image=update_image,bg="#38b6ff",command=update)
button2.place(x=20,y=20)
#menu
Registration=IntVar()
label3=Label(root,text="Registration NO:",font=("baskerville",15,"bold"),bg="#202329",fg="#ffffff")
label3.place(x=50,y=150)
entry2=Entry(root,width=15,bd=3,textvariable=Registration,font=("Baskerville",10,"bold"))
entry2.place(x=250,y=155)
registration_no()
Date=StringVar()
today=date.today()
d1=today.strftime("%d/%m/%y")
label4=Label(root,text="Date:",font=("baskerville",15,"bold"),bg="#202329",fg="#ffffff")
label4.place(x=450,y=150)
entry3=Entry(root,width=10,bd=3,text=Date,font=("Baskerville",10,"bold"))
entry3.place(x=540,y=155)
Date.set(d1)
#Student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg="#EDEDED",fg="#06283D",height=250,relief=GROOVE)
obj.place(x=50,y=200)
label5=Label(obj,text="Full Name: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label5.place(x=30,y=50)
label6=Label(obj,text="Date of Birth: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label6.place(x=30,y=100)
label7=Label(obj,text="Gender: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label7.place(x=30,y=150)
label8=Label(obj,text="Class: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label8.place(x=500,y=50)
label9=Label(obj,text="Religion: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label9.place(x=500,y=100)
label10=Label(obj,text="Skills: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label10.place(x=500,y=150)
Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="Baskerville")
name_entry.place(x=160,y=50)
DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=20,font="Baskerville")
dob_entry.place(x=160,y=100)
radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg="#EDEDED",fg="#06283D",command=selection)
R1.place(x=160,y=150)
R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg="#EDEDED",fg="#06283D",command=selection)
R2.place(x=250,y=150)
Class=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],width=17,font='Baskerville',state='r')
Class.place(x=630,y=50)
Class.set("Select Class")
Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,width=20,font="Baskerville")
religion_entry.place(x=630,y=100)
Skills=StringVar()
skills_entry=Entry(obj,textvariable=Skills,width=20,font="Baskerville")
skills_entry.place(x=630,y=150)
#Parent details
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg="#EDEDED",fg="#06283D",height=220,relief=GROOVE)
obj2.place(x=50,y=470)
label11=Label(obj2,text="Father's Name: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label11.place(x=30,y=50)
label12=Label(obj2,text="Occupation: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label12.place(x=30,y=100)
label13=Label(obj2,text="Mother's Name: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label13.place(x=500,y=50)
label14=Label(obj2,text="Occupation: ",font=("Baskerville",12,"bold"),bg="#EDEDED",fg="#06283D")
label14.place(x=500,y=100)
Father=StringVar()
father_entry=Entry(obj2,textvariable=Father,width=20,font="Baskerville")
father_entry.place(x=160,y=50)
fo=StringVar()
fo_entry=Entry(obj2,textvariable=fo,width=20,font="Baskerville")
fo_entry.place(x=160,y=100)
Mother=StringVar()
mother_entry=Entry(obj2,textvariable=Mother,width=20,font="Baskerville")
mother_entry.place(x=630,y=50)
mo=StringVar()
mo_entry=Entry(obj2,textvariable=mo,width=20,font="Baskerville")
mo_entry.place(x=630,y=100)
#Left Side
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)
img=PhotoImage(file='upload photo.png')
lbl=Label(f,bg='black',image=img)
lbl.place(x=0,y=0)
#button
Button(root,text="Upload",width=19,height=2,font=("Baskerville",12,'bold'),bg='lightblue',command=showimage).place(x=1000,y=370)
SaveButton=Button(root,text="Save",width=19,height=2,font=("Baskerville",12,'bold'),bg='lightgreen',command=save)
SaveButton.place(x=1000,y=450)
Button(root,text="Reset",width=19,height=2,font=("Baskerville",12,'bold'),bg='lightpink',command=clear).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font=("Baskerville",12,'bold'),bg='grey',command=exit).place(x=1000,y=610)
root.mainloop()